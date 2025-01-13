import pandas as pd
import fitz
import re
import os
from unidecode import unidecode 
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from PyPDF2 import PdfReader
from pathlib import Path

def extract_text_from_pdf(pdf_path):
    """Hàm đọc văn bản từ file PDF."""
    doc = fitz.open(pdf_path)
    output = []

    for page_number, page in enumerate(doc, start=1):
        blocks = page.get_text("blocks")
        for block in blocks:
            output.append({"page_number": page_number, "text": block[4]})
    
    return output

def read_keywords_from_txt(txt_path):
    """Hàm đọc từ khóa từ file txt."""
    with open(txt_path, 'r', encoding='utf-8') as file:
        content = file.read()
        # Tách từ khóa theo dấu ',' hoặc xuống dòng
        keywords = re.split(r'[,\n]+', content.strip())
    return [keyword.strip() for keyword in keywords if keyword.strip()]

def extract_sentence(pdf_text, keywords):
    """Trích xuất câu chứa từ khóa và thông tin số trang."""
    results = {}
    keywords_lower = [k.lower() for k in keywords]  # Chuyển từ khóa về dạng chữ thường trước để so sánh nhanh hơn.

    for item in pdf_text:
        # Tách văn bản thành các câu bằng biểu thức chính quy
        sentences = re.split(r'(?<=[.!?])\s+', item["text"])  # Tách tại dấu chấm, dấu chấm hỏi hoặc dấu chấm than, giữ khoảng trắng sau dấu câu
        page_number = item["page_number"]

        for sentence in sentences:
            sentence = sentence.strip()  # Loại bỏ khoảng trắng thừa
            # Kiểm tra nếu bất kỳ từ khóa nào xuất hiện trong câu (không phân biệt chữ hoa/thường)
            for keyword in keywords_lower:
                if keyword in sentence.lower():
                    if keyword not in results:
                        results[keyword] = []  # Khởi tạo danh sách nếu từ khóa chưa có trong kết quả
                    results[keyword].append(f"{sentence} ({page_number})")  # Thêm câu và số trang vào danh sách

    return results

def extract_paragraph(pdf_text, keywords):
    """Hàm tìm các đoạn văn chứa từ khóa."""
    results = {}
    for item in pdf_text:
        text = item["text"].strip()
        page_number = item["page_number"]

        for keyword in keywords:
            if keyword.lower() in text.lower():  # Luôn không phân biệt hoa-thường
                if keyword not in results:
                    results[keyword] = []  # Khởi tạo danh sách nếu từ khóa chưa có trong kết quả
                results[keyword].append({"paragraph": text, "page": page_number})  # Thêm đoạn văn và số trang
                break  # Dừng kiểm tra các từ khóa khác nếu đã tìm thấy
    
    return results

def clean_text_for_excel(text):
    """Loại bỏ các ký tự không hợp lệ trong Excel."""
    # Xóa các ký tự không hợp lệ (chẳng hạn như các ký tự có mã Unicode không hợp lệ cho Excel)
    text = re.sub(r'[^\x20-\x7E]', '', text)  # Loại bỏ ký tự ngoài phạm vi ASCII hợp lệ
    return text

def process_pdf_and_txt(txt_path, pdf_path, output_path):
    """Hàm chính để xử lý file txt và file PDF."""
    # Đọc từ khóa từ file txt
    keywords = read_keywords_from_txt(txt_path)

    # Đọc văn bản từ file PDF
    pdf_text = extract_text_from_pdf(pdf_path)

    # Trích xuất dữ liệu theo câu và đoạn văn
    found_segments_sentence = extract_sentence(pdf_text, keywords)  # Dự kiến trả về dict {"keyword": ["text_pagenumber", ...]}
    found_segments_paragraph = extract_paragraph(pdf_text, keywords)  # Dự kiến trả về dict {"keyword": ["text_pagenumber", ...]}

    # Tạo workbook mới hoặc mở workbook hiện tại
    wb = load_workbook(output_path) if Path(output_path).exists() else Workbook()

    # Xóa các sheet cũ nếu có
    for sheet_name in ['Sentence', 'Paragraph']:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    # Tạo sheet "Sentence"
    ws_sentence = wb.create_sheet("Sentence")
    ws_sentence.append(["Keyword", "Sentence", "Page Number"])  # Header
    for keyword, sentences in found_segments_sentence.items():
        for sentence in sentences:
            text, page_number = sentence.rsplit(" (", 1)  # Tách văn bản và số trang
            page_number = page_number.rstrip(")")  # Loại bỏ dấu ngoặc
            text = clean_text_for_excel(text)
            ws_sentence.append([keyword, text, page_number])  # Thêm dữ liệu vào sheet "Sentence"

    # Tạo sheet "Paragraph"
    ws_paragraph = wb.create_sheet("Paragraph")
    ws_paragraph.append(["Keyword", "Paragraph", "Page Number"])  # Header
    for keyword, paragraphs in found_segments_paragraph.items():
        for paragraph in paragraphs:
            text = paragraph["paragraph"]  # Tách văn bản và số trang
            page_number = paragraph["page"]  # Loại bỏ dấu ngoặc
            text = clean_text_for_excel(text)
            ws_paragraph.append([keyword, text, page_number])  # Thêm dữ liệu vào sheet "Paragraph"

    # Xóa sheet mặc định nếu không dùng
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    # Căn chỉnh Excel (giả sử đã có hàm adjust_excel_format)
    adjust_excel_format(wb)

    # Lưu file Excel
    wb.save(output_path)
    print(f"Kết quả đã được lưu vào {output_path}")


def adjust_excel_format(wb):
    """Hàm căn chỉnh nội dung và định dạng cho sheet Excel."""
    for sheet_name in ["Sentence", "Paragraph"]:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):  # Chỉ căn chỉnh nội dung văn bản
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Căn chỉnh độ rộng cột B
        col_letter = 'B'
        max_length = 0
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

def process_directory(pdf_directory, txt_path, output_directory):
    """Hàm đọc tất cả các file PDF trong thư mục và xuất dữ liệu vào các file Excel."""
    # Đảm bảo thư mục đầu ra tồn tại
    Path(output_directory).mkdir(parents=True, exist_ok=True)

    # Duyệt qua tất cả các file PDF trong thư mục
    for pdf_file in os.listdir(pdf_directory):
        if pdf_file.endswith(".pdf"):
            pdf_path = os.path.join(pdf_directory, pdf_file)
            output_file = f"{os.path.splitext(pdf_file)[0]}.xlsx"  # Đổi đuôi .pdf thành .xlsx
            output_path = os.path.join(output_directory, output_file)

            # Xử lý PDF và lưu vào Excel
            process_pdf_and_txt(txt_path, pdf_path, output_path)

txt_path = 'keywords.txt' 
pdf_dir = 'file_pdf'  
output_dir = 'excel'
output_path = 'output.xlsx'
process_directory(pdf_dir, txt_path, output_dir)
