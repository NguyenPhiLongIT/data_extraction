import os  
import fitz
from openpyxl import load_workbook
import re


def remove_header_footer(pdf_extracted_text):
    page_format_pattern = r'([pP]+[aA]+[gG]+[\s]*[\d]+)'
    url_pattern = r'https?://[^\s]+'
    pdf_extracted_text = pdf_extracted_text.split("\n")
    header = pdf_extracted_text[0].strip()
    footer = pdf_extracted_text[-1].strip()
    
    if re.search(page_format_pattern, header) or header.isnumeric():
        pdf_extracted_text = pdf_extracted_text[1:]
    
    if re.search(page_format_pattern, footer) or footer.isnumeric():
        pdf_extracted_text = pdf_extracted_text[:-1]
    
    pdf_extracted_text = [re.sub(url_pattern, '', line) for line in pdf_extracted_text]
    pdf_extracted_text = "\n".join(pdf_extracted_text)
    return pdf_extracted_text

def extract_pdf_excel(input_pdf, output_excel):
    try:
        if not os.path.isfile(input_pdf):
            print(f"Tệp PDF không tồn tại: {input_pdf}")
            return

        wb = load_workbook(output_excel)
        ws = wb.active

        # Đọc danh sách từ khóa từ cột D
        keywords = []
        for row in ws.iter_rows(min_col=4, max_col=4, min_row=1, values_only=True):
            if row[0]:
                keywords.append(row[0])

        if not keywords:
            print("Không có từ khóa nào được tìm thấy trong cột D.")
            return

        print(f"Danh sách từ khóa: {keywords}")

        pdf_document = fitz.open(input_pdf)
        num_pages = pdf_document.page_count
        print(f"Số trang trong PDF: {num_pages}")

        sentence_count = 2  # Bắt đầu ghi từ dòng 2 trong Excel
        for page_number in range(7, min(345, num_pages)):  # Giới hạn từ trang 7 đến 345
            page = pdf_document[page_number]
            text = page.get_text("text")

            if not text:
                print(f"Trang {page_number + 1} không có nội dung.")
                continue

            keyword_match_count = 0
            keyword_match = []
            for keyword in keywords:
                if keyword.lower() in text.lower():
                    keyword_match_count += 1
                    keyword_match.append(keyword)

            # Nếu tìm thấy từ khóa, lưu thông tin vào Excel
            if keyword_match_count > 0:
                text = remove_header_footer(text)

                ws[f"A{sentence_count}"] = text
                ws[f"B{sentence_count}"] = ", ".join(keyword_match)
                ws[f"C{sentence_count}"] = page_number + 1
                sentence_count += 1
                print(f"Trang {page_number + 1}: Tìm thấy {keyword_match}")

        wb.save(output_excel)
        print(f"Kết quả đã được lưu vào: {output_excel}")

        pdf_document.close()

    except Exception as e:
        print(f"Đã xảy ra lỗi: {e}")

if __name__ == "__main__":
    input_pdf = "file/the-fifth-discipline-the-art-and-practice-of-the-learning-organization_compress.pdf"
    output_excel = "file/output.xlsx"

    extract_pdf_excel(input_pdf, output_excel)