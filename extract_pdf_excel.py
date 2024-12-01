import os
from openpyxl.styles import Font
import fitz
from openpyxl import load_workbook
import re
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl import Workbook
import xlsxwriter
from openpyxl.styles import Alignment

from openpyxl.utils import get_column_letter
def remove_extra_spaces(text):
    return " ".join(text.split())
def adjust_column_width(ws, col):
    """Điều chỉnh độ rộng cột cho phù hợp với nội dung"""
    max_length = 0
    for row in ws.iter_rows(min_col=col, max_col=col):
        for cell in row:
            try:
                # Tính độ dài của văn bản trong ô
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
            except:
                pass
    adjusted_width = (max_length + 2)  # Thêm một ít không gian để nhìn thoáng hơn
    ws.column_dimensions[get_column_letter(col)].width = adjusted_width

def adjust_row_height(ws, row):
    """Điều chỉnh chiều cao dòng tự động sao cho phù hợp với nội dung"""
    min_height = 30  # Chiều cao tối thiểu cho một dòng
    line_height = 30  # Chiều cao của một dòng (có thể điều chỉnh tùy vào font và kích thước chữ)

    max_height = min_height  # Khởi tạo chiều cao tối thiểu

    # Duyệt qua các ô trong dòng
    for cell in ws[row]:
        try:
            # Lấy giá trị của ô và chuyển thành chuỗi
            text = str(cell.value)
            # Đếm số dòng (dựa vào ký tự ngắt dòng '\n')
            num_lines = text.count('\n') + 1  # Cộng thêm 1 dòng để tính dòng cuối cùng

            print(f"Số dòng trong ô {cell.coordinate}: {num_lines}")  # In số dòng để kiểm tra

            # Tính chiều cao cần thiết cho ô (số dòng nhân với chiều cao của một dòng)
            max_height = max(max_height, num_lines * line_height)

            # Cài đặt thuộc tính wrap_text cho ô để tự động xuống dòng
            cell.alignment = Alignment(wrap_text=True)

        except:
            pass

    # Đảm bảo chiều cao dòng không thấp hơn chiều cao tối thiểu
    ws.row_dimensions[row].height = max(max_height, min_height)
def remove_header_footer(pdf_extracted_text):
    page_format_pattern = r'([pP]+[aA]+[gG]+[\s]*[\d]+)'
    url_pattern = r'https?://[^\s]+'
    pdf_extracted_text = pdf_extracted_text.split("\n")
    header = pdf_extracted_text[0].strip()
    footer = pdf_extracted_text[-1].strip()
    
    if re.search(page_format_pattern, header) or header.isnumeric():
        pdf_extracted_text = pdf_extracted_text[1:]
    
    if re.search(page_format_pattern, footer) or footer.isnumeric():
        pdf_extracted_text = pdf_extracted_text[:-1]
    
    pdf_extracted_text = [re.sub(url_pattern, '', line) for line in pdf_extracted_text]
    pdf_extracted_text = "\n".join(pdf_extracted_text)
    return pdf_extracted_text

def extract_pdf_excel(input_pdf, output_excel):
    try:
        if not os.path.isfile(input_pdf):
            print(f"Tệp PDF không tồn tại: {input_pdf}")
            return

        # Mở file Excel có sẵn để đọc từ khóa
        wb = load_workbook(output_excel)
        ws = wb.active
        highlight_font = Font(color="FF0000", bold=True)  # Màu đỏ và in đậm
        normal_font = Font(color="000000")  # Màu đen

        # Đọc danh sách từ khóa từ cột D (cột thứ 4) của worksheet
        keywords = []
        for row in ws.iter_rows(min_col=4, max_col=4, min_row=1, values_only=True):
            if row[0]:
                keywords.append(row[0])

        if not keywords:
            print("Không có từ khóa nào được tìm thấy trong cột D.")
            return

        print(f"Danh sách từ khóa: {keywords}")

        # Mở tài liệu PDF
        pdf_document = fitz.open(input_pdf)
        num_pages = pdf_document.page_count
        print(f"Số trang trong PDF: {num_pages}")

        sentence_count = 2  # Bắt đầu ghi từ dòng 2 trong Excel
        for page_number in range(7, min(345, num_pages)):  # Giới hạn từ trang 7 đến 345
            print("page_number", page_number)
            page = pdf_document[page_number]

            text = page.get_text("text")

            if not text:
                print(f"Trang {page_number + 1} không có nội dung.")
                continue

            text = remove_extra_spaces(text)
            for sentence in text.split("."):  # Chia văn bản thành các câu
                sentence = sentence.strip()
                if not sentence:
                    continue

                keyword_match_count = 0
                keyword_match = []
                highlighted_sentence = sentence  # Mặc định là câu không có highlight

                # Lặp qua từng từ khóa
                for keyword in keywords:
                    if keyword.lower() in sentence.lower():  # Kiểm tra từ khóa có trong câu
                        keyword_match_count += 1
                        keyword_match.append(keyword)
                        # Highlight từ khóa trong câu bằng cách thay thế
                        highlighted_sentence = highlighted_sentence.replace(" "+keyword, f"{{{{{keyword}}}}}")
                        

                # Nếu tìm thấy từ khóa trong câu, lưu thông tin vào Excel
                if keyword_match_count > 0:
                    # Ghi câu vào cột A
                    ws[f"A{sentence_count}"] = highlighted_sentence
                    ws[f"A{sentence_count}"].alignment = Alignment(wrap_text=True)
                    # Ghi từ khóa vào cột B
                    ws[f"B{sentence_count}"] = ", ".join(keyword_match)

                    # Ghi số trang vào cột C
                    ws[f"C{sentence_count}"] = page_number+1
                    # adjust_column_width(ws, 1)  # Điều chỉnh độ rộng cột A
                    adjust_row_height(ws, sentence_count) 
                    sentence_count += 1
                    print(f"Trang {page_number+1}: Tìm thấy {keyword_match}")

        # Lưu kết quả vào Excel
        wb.save(output_excel)
        print(f"Kết quả đã được lưu vào: {output_excel}")

        pdf_document.close()

    except Exception as e:
        print(f"Đã xảy ra lỗi: {e}")

if __name__ == "__main__":
    input_pdf = "output1.pdf"
    output_excel = "file/output1.xlsx"
    # keywords=['chúng ta', "ta có","ta","có", "biết","như","xác định", "đồ thị","truy vấn","dữ liệu"]
    extract_pdf_excel(input_pdf, output_excel)